from abc import ABCMeta, abstractmethod
import NER.util.ner_configs
from NER.util.ner_configs import person_config
import datetime
from datetime import timedelta
import logging
import openpyxl
import os
from NER.vendors.vendor import ChargeType



class Spreadsheet(metaclass=ABCMeta):

    def __init__(self, template_filename, output_filename):
        """
        Initializes a spreadsheet class

        @:arg template_filename: the name of a template spreadsheet (must exist in the directory of the class
        @:arg output_filename: the name of the file to write out (can be absolute, otherwise relative to cwd
        """
        self.logger = logging.getLogger(__name__)
        self._vendors = []
        self._template_filename = template_filename
        self._output_filename = output_filename
        self._workbook = None
        self.__from_date = None
        self.__to_date = None

    def create_and_open_output_file(self):
        template_full_path = os.path.join(os.path.dirname(__file__), self._template_filename)
        self._workbook = openpyxl.load_workbook(template_full_path)
        self._workbook.save(self._output_filename)

    def process_static_information(self):
        self._employee_name_cell.value = person_config['name']
        self._employee_number_cell.value = person_config['employee_number']
        self._hor_cell.value = person_config['home_address']
        self._date_prepared_cell.value = datetime.datetime.now().strftime("%Y-%m-%d")
        self._travel_destination_cell.value = self._travel_destination
        self._lodging_per_diem_cell.value = self._lodging_per_diem
        self._meal_per_diem_cell.value = self._meal_per_diem

        self._employee_signed_name_cell.value = person_config['name']
        self._employee_signed_date_cell.value = datetime.datetime.now().strftime("%Y-%m-%d")
        self._approver_signed_name_cell.value = person_config['employee_approver']
        self._workbook.save(self._output_filename)

    def process_date_information(self):
        self._to_date_cell.value = self._to_date
        self._from_date_cell.value = self._from_date

        number_of_days = (self._to_date - self.__from_date).days
        if (number_of_days > 7):
            raise NotImplementedError("This program cannot handle expense reports for a period over 7 days")

        # Set the title and per diem
        first_day_title_cell = self._first_day_title_cell
        for i in range(number_of_days + 1):
            title_cell = self._sheet.cell(first_day_title_cell.row, first_day_title_cell.column + i)
            title_date = self.__from_date + timedelta(days=i)
            title_cell.value = title_date

            per_diem_cell = self.get_cell_for_date_and_charge(date=title_date, charge_type=ChargeType.MEAL_PER_DIEM)
            per_diem_cell.value = self._meal_per_diem
            if title_date == self._from_date or title_date == self._to_date:
                per_diem_cell.value = self._meal_per_diem * .75

        self._workbook.save(self._output_filename)

    def process_vendors(self):
        for vendor in self._vendors:
            for charge in vendor.get_expense_charges():
                cell = self.get_cell_for_date_and_charge(charge.charge_date, charge.charge_type)

                prev_value = cell.value
                if prev_value is None:
                    prev_value = 0
                cell.value = prev_value + charge.charge_amount

        self._workbook.save(self._output_filename)

    @property
    @abstractmethod
    def _hor_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _employee_name_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _employee_number_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _date_prepared_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _travel_destination_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _travel_destination(self):
        """@:return string with the name of the travel destination"""
        raise NotImplementedError

    @property
    @abstractmethod
    def _lodging_per_diem(self):
        """
        :return: number of the daily lodging per diem
        """
        raise NotImplementedError

    @property
    @abstractmethod
    def _lodging_per_diem_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _meal_per_diem(self):
        """
        :return: number of the daily lodging per diem
        """
        raise NotImplementedError

    @property
    @abstractmethod
    def _meal_per_diem_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _employee_signed_name_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _employee_signed_date_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _approver_signed_name_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _from_date_cell(self):
        raise NotImplementedError

    @property
    @abstractmethod
    def _to_date_cell(self):
        raise NotImplementedError

    def __get_all_dates(self):
        all_dates = []
        for a_list_of_list in [v.get_expense_dates() for v in self._vendors]:
            for i in range(len(a_list_of_list)):
                all_dates.append(a_list_of_list[i])
        return all_dates

    @property
    def _from_date(self):
        if self.__from_date is None:
            self.__from_date = min(self.__get_all_dates())
        return self.__from_date

    @property
    def _to_date(self):
        if self.__to_date is None:
            self.__to_date = self.__to_date = max(self.__get_all_dates())
        return self.__to_date

    @property
    @abstractmethod
    def _sheet(self):
        """@:return the sheet with the expense report"""
        raise NotImplementedError

    @property
    @abstractmethod
    def _first_day_title_cell(self):
        """
        :return: The cell for the title of the first day.

        This program assumes a day is handled in a column. And the columns are sequential
        """
        raise NotImplementedError

    @abstractmethod
    def get_cell_for_date_and_charge(self, date, charge_type):
        """ :return: The cell for charge_type on date

        @param date datetime of the desired cell
        @param charge_type ChargeType of the desired cell
        """
        raise NotImplementedError

    def set_vendors(self, vendors):
        self._vendors = vendors

    def add_vendor(self, new_vendor):
        self._vendors.append(new_vendor)

    def get_cell_home_of_record(self):
        return self._hor_cell


class SanAntonio2020SpreadSheet(Spreadsheet):

    @property
    def _employee_name_cell(self):
        return self._workbook['Travel Expense']['C5']

    @property
    def _employee_number_cell(self):
        return self._workbook['Travel Expense']["G5"]

    @property
    def _date_prepared_cell(self):
        return self._workbook['Travel Expense']["J5"]

    @property
    def _hor_cell(self):
        return self._workbook['Travel Expense']["D10"]

    @property
    def _travel_destination_cell(self):
        return self._workbook['Travel Expense']["D11"]

    @property
    def _travel_destination(self):
        return "San Antonio"

    @property
    def _lodging_per_diem_cell(self):
        return self._workbook['Travel Expense']["D12"]

    @property
    def _lodging_per_diem(self):
        return 126

    @property
    def _meal_per_diem_cell(self):
        return self._workbook['Travel Expense']["D13"]

    @property
    def _meal_per_diem(self):
        return 61

    @property
    def _employee_signed_name_cell(self):
        return self._workbook['Travel Expense']["C40"]

    @property
    def _employee_signed_date_cell(self):
        return self._workbook['Travel Expense']["K40"]

    @property
    def _approver_signed_name_cell(self):
        return self._workbook['Travel Expense']["C42"]

    @property
    def _from_date_cell(self):
        return self._workbook['Travel Expense']["D9"]

    @property
    def _to_date_cell(self):
        return self._workbook['Travel Expense']["E9"]

    @property
    def _first_day_title_cell(self):
        return self._workbook['Travel Expense']["D17"]

    def get_cell_for_date_and_charge(self, date, charge_type):
        col_delta_from_first_date = (date - self._from_date).days
        row_delta_from_title = {ChargeType.AIRPLANE: 2,
                                ChargeType.UBER: 3,
                                ChargeType.PARKING: 4,
                                ChargeType.HOTEL: 5,
                                ChargeType.LODGING_TAX: 7,
                                ChargeType.MEAL_PER_DIEM: 9,
                                ChargeType.PHONE: 10,
                                ChargeType.CAR_RENTAL: 11,
                                ChargeType.GAS: 12}[charge_type]
        first_day_title_cell = self._first_day_title_cell
        return self._sheet.cell(first_day_title_cell.row + row_delta_from_title,
                                first_day_title_cell.column + col_delta_from_first_date)

    @property
    def _sheet(self):
        return self._workbook['Travel Expense']

    def __init__(self, template_filename='2020_sa_template.xlsx', output_filename='{date}_{user}_expense_report.xls'):
        super(SanAntonio2020SpreadSheet, self).__init__(template_filename=template_filename,
                                                        output_filename=output_filename)
